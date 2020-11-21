import ui_Excel_ui

from PyQt5.QtWidgets import QApplication, QSplashScreen, QWidget, QMessageBox, QTableWidgetItem, QAbstractItemView
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QThreadPool
from PyQt5.QtGui import QBrush, QPixmap, QIcon

from re import compile
from os.path import isfile, isdir
import wget
from time import sleep
import os
import openpyxl
import sys


class DL_file:
    """
    下载对象的类
    对每个下载对象会对应一个本类
    包含功能：
    下载：Download
    """

    def __init__(self):
        self.name = '未命名'
        self.num = 0
        self.url = []
        self.ff = []

    def Download(self, save_path, md_dir: bool) -> bool:
        """
        类的下载模块，可以对类的链接进行下载和重命名
        :param save_path:
        :param md_dir: 是否新建文件夹
        :return: 无
        """
        if md_dir and not os.path.isdir(save_path + '\\' + self.name):
            os.mkdir(save_path + '\\' + self.name)
        print(self.name)
        if self.num == 1:
            num_temp = 2
            temp = True
            if md_dir:
                filename = save_path + '\\' + self.name + '\\' + self.name + self.ff[0]
            else:
                filename = save_path + '\\' + self.name + self.ff[0]
            while temp:
                if os.path.isfile(filename):
                    if md_dir:
                        filename = save_path + '\\' + self.name + '\\' \
                                   + self.name + '(' + str(num_temp) + ')' + self.ff[0]
                        num_temp += 1
                    else:
                        filename = save_path + '\\' + self.name + '(' + str(num_temp) + ')' + self.ff[0]
                        num_temp += 1
                else:
                    temp = False
            temp = True
            while temp:
                try:
                    wget.download(self.url[0], filename)
                except ConnectionResetError:
                    sleep(0.5)
                    continue
                temp = False
        elif self.num > 1:
            num_temp = 1
            for i in range(0, self.num):
                temp = True
                while temp:
                    if md_dir:
                        filename = save_path + '\\' + self.name \
                                   + '\\' + self.name + '(' + str(num_temp) + ')' + self.ff[i]
                    else:
                        filename = save_path + '\\' + self.name + '(' + str(num_temp) + ')' + self.ff[i]
                    if os.path.isfile(filename):
                        num_temp += 1
                    else:
                        temp = False
                temp = True
                while temp:
                    try:
                        if md_dir:
                            wget.download(self.url[i], save_path + '\\' + self.name + '\\' +
                                          self.name + '(' + str(num_temp) + ')' + self.ff[i])
                        else:
                            wget.download(self.url[i], save_path + '\\' +
                                          self.name + '(' + str(num_temp) + ')' + self.ff[i])
                    except ConnectionResetError:
                        sleep(0.5)
                    temp = False
        return True


class Downloading(QThread):
    """
    下载线程的类，基于QThread
    用于对单个对象的下载
    """
    FinishSignal = pyqtSignal()

    def __init__(self, DLfile, mk, Save_path):
        super(Downloading, self).__init__()
        self.DLfile = DLfile
        self.mk = mk
        self.Save_path = Save_path

    def run(self) -> None:
        self.DLfile.Download(self.Save_path, self.mk)
        self.FinishSignal.emit()


def DlMax(dl):
    num_list = []
    for i in dl:
        num_list.append(i.num)
    # return 1
    return max(num_list)


def fileInput(file_path: str, column_name: int, DL: list) -> None:
    """
    用以打开xlsx文件并提取文件名和链接地址
    :param DL: 存储类的列表
    :param file_path:文件地址
    :param column_name:重命名所在列号
    :return:无
    """
    file = openpyxl.load_workbook(file_path)  # 打开Excel文件
    sheet_name = file.sheetnames  # 获取第一个表的名字
    sheet = file.get_sheet_by_name(sheet_name[0])  # 读取第一个表
    re_url = compile('^(https://)')  # 判断是否是网址
    for x in range(2, sheet.max_row + 1):
        if sheet.cell(row=x, column=column_name).value is None:
            continue
        temp = DL_file()
        temp.name = sheet.cell(row=x, column=column_name).value
        i = 1
        while not sheet.cell(row=x, column=column_name + i).value is None:
            if re_url.match(str(sheet.cell(row=x, column=column_name + i).value)):
                text = sheet.cell(row=x, column=column_name + i).value
            else:
                try:
                    text = sheet.cell(row=x, column=column_name + i).hyperlink.target
                except AttributeError:
                    i += 1
                    continue
            if text is None:
                i += 1
                continue
            temp.url.append(text)
            temp.ff.append(r'.' + text[text.index(r'type') + 5:])
            temp.num += 1
            i += 1
        DL.append(temp)


class runDownload(QThread):
    finishDownload_signal = pyqtSignal()
    DownloadNums_signal = pyqtSignal(int)

    def __init__(self, gui):
        super(QThread, self).__init__()
        self.dl = gui.DL
        self.num = len(self.dl)
        self.save_path = gui.ui.SaveTextEdit.toPlainText()
        self.md_dir = gui.ui.checkBox.isChecked()
        self.FinishedNums = 0
        self.ThreadNums = int(gui.ui.ThreadsSpinBox.text())
        self.ThreadsPool = QThreadPool()
        self.ThreadsPool.setMaxThreadCount(int(gui.ui.ThreadsSpinBox.text()))

    def run(self) -> None:
        runList = []
        for temp in self.dl:
            while True:
                if len(runList) < self.ThreadNums:
                    self.inputThreadList(runList, temp)
                    break
                else:
                    for i in runList:
                        if i.isFinished():
                            runList.remove(i)
        while len(runList) > 0:
            for i in runList:
                if i.isFinished():
                    runList.remove(i)

        self.finishDownload_signal.emit()
        self.RemoveTempFile()
        print('start ' + '\'' + self.save_path + '\'')
        os.startfile(self.save_path)

    def inputThreadList(self, runList, dl):
        down = Downloading(dl, self.md_dir, self.save_path)
        runList.append(down)
        down.FinishSignal.connect(self.ed_signal)
        down.start()

    def RemoveTempFile(self):
        if self.md_dir:
            a = os.listdir(self.save_path)
            for j in a:
                b = os.listdir(self.save_path + '/' + j)
                path_temp = self.save_path + '/' + j
                for i in b:
                    if i[-3:] == 'tmp':
                        os.remove(path_temp + r'/' + i)
        else:
            a = os.listdir(self.save_path)
            for i in a:
                if i[-3:] == 'tmp':
                    os.remove(self.save_path + '/' + i)

    def ed_signal(self):
        self.FinishedNums += 1
        self.DownloadNums_signal.emit(int((self.FinishedNums / self.num) * 100))


class QmyWidget(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = ui_Excel_ui.Ui_Excel_ui()
        self.ui.setupUi(self)
        self.ui.TableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.DL = []
        self.runDownload = None

    def on_StartPushButton_clicked(self):
        self.ui.StartPushButton.setEnabled(False)
        file_path = self.ui.FileTextEdit.toPlainText()
        save_path = self.ui.SaveTextEdit.toPlainText()
        temp1 = compile(r'^.+\.(xlsx|xls)$')
        if file_path == '':
            QMessageBox.warning(self, '输入错误！', '请输入文件地址！')
            print('\a')
        elif not (temp1.match(file_path) and isfile(file_path)):
            QMessageBox.about(self, '输入错误！', '文件地址输入错误，只支持xls/xlsx文件，请重新输入！')
            print('\a')
            self.ui.FileTextEdit.clear()
        elif save_path == '':
            QMessageBox.about(self, '输入错误！', '请输入文件夹地址！')
            print('\a')
        elif not isdir(save_path):
            QMessageBox.about(self, '输入错误！', '地址输入错误或文件夹不存在，请重新输入输入文件夹地址！')
            print('\a')
            self.ui.SaveTextEdit.clear()
        else:
            if QMessageBox.question(self, '开始下载', '是否开始下载？\n下载完成后会自动打开文件所在目录，请耐心等待',
                                    QMessageBox.Ok | QMessageBox.Cancel) == QMessageBox.Ok:
                self.ui.progressBar.setEnabled(True)
                self.ui.progressBar.setValue(0)
                # self.ui.StartPushButton.setEnabled(False)
                self.ui.StartPushButton.setText('下载中...')
                self.runDownload = runDownload(self)
                self.runDownload.DownloadNums_signal.connect(self.ChangeProgressBar)
                self.runDownload.finishDownload_signal.connect(self.FinshDownload)
                self.runDownload.start()
                self.runDownload.exec()
            else:
                self.ui.StartPushButton.setEnabled(True)
        self.ui.StartPushButton.setEnabled(True)

    def on_SaveTextEdit_textChanged(self):
        if 0 == self.ui.SaveTextEdit.toPlainText().find('file:///'):
            self.ui.SaveTextEdit.setText(self.ui.SaveTextEdit.toPlainText().replace('file:///', ''))

    def on_FileTextEdit_textChanged(self):
        temp1 = compile(r'^.+\.(xlsx|xls)$')
        if 0 == self.ui.FileTextEdit.toPlainText().find('file:///'):
            self.ui.FileTextEdit.setText(self.ui.FileTextEdit.toPlainText().replace('file:///', ''))
        if temp1.match(self.ui.FileTextEdit.toPlainText()) and isfile(self.ui.FileTextEdit.toPlainText()):
            self.TableView()

    def on_ColumnSpinBox_valueChanged(self):
        self.on_FileTextEdit_textChanged()

    def TableView(self):
        self.DL.clear()
        self.ui.TableWidget.clearContents()
        file_path = self.ui.FileTextEdit.toPlainText()
        column_name = self.ui.ColumnSpinBox.text()
        column_name = int(column_name)
        fileInput(file_path, column_name, self.DL)
        if len(self.DL) == 0:
            QMessageBox.warning(self, '读取失败', '读取失败，请检查调整命名列号，或检查文件！')
            print('\a')
            return
        ColumnCount = DlMax(self.DL)
        self.ui.TableWidget.setColumnCount(ColumnCount + 2)

        headerItem = QTableWidgetItem('文件名')
        font = headerItem.font()
        font.setPointSize(11)
        headerItem.setFont(font)
        headerItem.setForeground(QBrush(Qt.black))  # 前景色，即文字颜色
        self.ui.TableWidget.setHorizontalHeaderItem(0, headerItem)

        headerItem = QTableWidgetItem('文件数量')
        font = headerItem.font()
        font.setPointSize(11)
        headerItem.setFont(font)
        headerItem.setForeground(QBrush(Qt.black))  # 前景色，即文字颜色
        self.ui.TableWidget.setHorizontalHeaderItem(1, headerItem)

        for i in range(ColumnCount):
            headerItem = QTableWidgetItem('下载地址' + str(i + 1))
            font = headerItem.font()
            font.setPointSize(11)
            headerItem.setFont(font)
            headerItem.setForeground(QBrush(Qt.black))  # 前景色，即文字颜色
            self.ui.TableWidget.setHorizontalHeaderItem(i + 2, headerItem)

        self.ui.TableWidget.setRowCount(len(self.DL))
        self.ui.TableWidget.setAlternatingRowColors(True)  # 设置交替行背景颜色
        for i in range(len(self.DL)):
            item = QTableWidgetItem()
            item.setText(str(self.DL[i].name))
            self.ui.TableWidget.setItem(i, 0, item)
            item = QTableWidgetItem()
            item.setText(str(self.DL[i].num))
            self.ui.TableWidget.setItem(i, 1, item)
            for j in range(self.DL[i].num):
                item = QTableWidgetItem()
                item.setText(self.DL[i].url[j])
                self.ui.TableWidget.setItem(i, j + 2, item)

    def ChangeProgressBar(self, n):
        self.ui.progressBar.setValue(n)

    def FinshDownload(self):
        print('\a')
        QMessageBox.about(self, '下载完成！', '下载已完成！')
        self.ui.StartPushButton.setEnabled(True)
        self.ui.StartPushButton.setText('开始下载')


if __name__ == '__main__':
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    app = QApplication(sys.argv)
    # 创建启动界面
    splash = QSplashScreen(QPixmap('logo.ico'))
    splash.show()
    mainWindow = QWidget()
    ui = QmyWidget(mainWindow)
    mainWindow.setWindowTitle('Excel批量下载')
    mainWindow.setWindowIcon(QIcon('logo.ico'))
    mainWindow.show()
    splash.close()
    sys.exit(app.exec_())
